import openpyxl
import openai
from tqdm import tqdm
import threading
import pickle
from dotenv import load_dotenv
import os


def formatter(string):
    split = string.split()
    keys = []
    for i, item in enumerate(split):
        if "@" in item:
            keys.append(item.replace(
                "@", "").replace("[", "").replace("]", ""))
            split[i] = "{}"
    return (" ".join(split), keys)


def GPTrequest(prompt="Qual è il fatturato di amazon?"):
    return openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "user", "content": prompt},
        ]
    )["choices"][0]["message"]["content"]


def generateStint(len, num_thread):
    if len == 0:
        return []
    if len == 1:
        return [(0, 1)]
    stints = []
    step = round(len / num_thread)
    for i in range(2, len, step):
        end = i+step if i+step < len else len
        stints.append((i, end))
    return stints


def thread_func(content, header, start_index, end_index, new_fields, result, index):
    loc_result = []
    for i in range(start_index, end_index):
        if (end_index - i) % 10 == 0:
            print(f'Thread {index}: index:{i} - left: {end_index-i}')
        loc_content = {}
        for j in range(0, len(header)):
            loc_content[header[j]] = content[i][header[j]]
        for k in list(new_fields.keys()):
            flag = True
            while (flag):
                try:
                    loc_content[k] = GPTrequest(new_fields[k][0].format(
                        *[loc_content[x] for x in new_fields[k][1]]))
                    flag = False
                except:
                    pass
        loc_result.append(loc_content)
        with open(f'partial_{index}.pickle', 'wb') as fp:
            pickle.dump(loc_result, fp, protocol=pickle.HIGHEST_PROTOCOL)
    result[index] = loc_result
    save(f"./output/aziende_output_{index}.xlsx", loc_result)


def save(path, content):
    print("Saving...")
    wb = openpyxl.Workbook()
    sheet = wb.active

    for i, k in enumerate(list(content[0].keys())):
        sheet.cell(row=1, column=i+1).value = k

    for i, item in enumerate(content):
        for j, k in enumerate(list(item.keys())):
            sheet.cell(row=i+2, column=j+1).value = content[i][k]
    wb.save(path)


def load(path):
    print("Loading file...")
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column

    header = []
    for i in range(1, m_col + 1):
        header.append(sheet_obj.cell(row=1, column=i).value)

    content = []
    for i in tqdm(range(2, m_row + 1)):
        loc_content = {}
        for j in range(1, m_col + 1):
            loc_content[header[j-1]] = sheet_obj.cell(row=i, column=j).value
        content.append(loc_content)
    content = list(
        filter(lambda item: item["Customer Name"] is not None, content))
    return (header, content)




load_dotenv()

# CONFIGURATION
chat_gpt_api_key = os.environ["API_KEY"]
input_path = os.environ["INPUT_PATH"]
output_path = os.environ["OUTPUT_PATH"]
num_threads = os.environ["NUM_THREADS"]

openai.api_key = chat_gpt_api_key

input_key_question = {
    "Storia": "In accordo con le informazioni presenti sul web. Di cosa si occupa l'azienda @[Customer Name]@",
    "Fatturato": "In accordo con le informazioni presenti sul web. Qual è il fatturato dell'azienda @[Customer Name]@"
}

key_question_map = {}
for k in list(key_question_map.keys()):
    key_question_map[k] = formatter(input_key_question[k])

header, content = load(input_path)

stints = generateStint(len(content), num_threads)
results = [None] * num_threads

threads = []

for i, (start, end) in enumerate(stints):
    t = threading.Thread(target=thread_func,
                         args=(content, header, start, end, key_question_map, results, i))
    threads.append(t)
    print("Main    : starting thread %d.", i)
    t.start()

for index, thread in enumerate(threads):
    print("Main    : before joining thread %d.", index)
    thread.join()
    print("Main    : thread %d done", index)

with open('results.pickle', 'wb') as fp:
    pickle.dump(results, fp, protocol=pickle.HIGHEST_PROTOCOL)

result = [j for sub in results for j in sub]

save(output_path, result)
